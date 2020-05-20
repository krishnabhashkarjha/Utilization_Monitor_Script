"""
@ Author - Krishnabhashkar jha
@ Creation date - 15/01/2020
@ Description - All Data validation code.
"""

import openpyxl
from openpyxl import load_workbook
class ExcelOperations :

    def search_for_element_in_column(self, column_index, key):
        wb = openpyxl.load_workbook(self)
        sheet = wb.active
        for i in range(1, sheet.max_row1+1):
            temp = sheet.cell(row=i, column = column_index).value
            if temp == key:
                return i
        return 0

    def get_value(file_name, row_index, column_index):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        v_value = sheet.cell(row = row_index, column = column_index)
        return  v_value.value

    def set_value_to_cell(file_name, row_index, column_index, key):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        sheet.cell(row_index, column_index, key)
        wb.save(filename=file_name)

    def Excel_Update(self,row,col):
        v_filePath = "C:\Automation\Results.xlsx"
        v_row = row
        v_col = col
        status = 'Successfull'
        ExcelOperations.set_value_to_cell(v_filePath, v_row, v_col, status)

    def delete_first_row(file_name):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        sheet.delete_rows(sheet.min_row, 1)
        wb.save(file_name)



