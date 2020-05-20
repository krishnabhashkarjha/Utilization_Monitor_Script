"""
@ Author - Krishnabhashkar jha
@ Creation date - 26/03/2020
@ Description - Declares all the Methods to be used at the Process Level.
"""


import time
import openpyxl
from Applications.Workflows.UtilizationCalculator.TogglUtilizationCalculator import ElementLocators as Element
from Applications.Workflows.UtilizationCalculator.Utilites.SeleniumOperations import SeleniumOperations
from Applications.Workflows.UtilizationCalculator.TogglUtilizationCalculator import UserCredentials
from Applications.Workflows.UtilizationCalculator.Utilites.ExcelOperations import ExcelOperations
class UtilizationMethod:

    def __init__(self,v_Browser,lo):
        self.v_Browser = v_Browser
        self.lo = lo

    def Login(self):
        try:
            SeleniumOperation = SeleniumOperations(self.v_Browser, self.lo)
            self.v_Browser.get('https://toggl.com/login/')
            time.sleep(5)
            SeleniumOperation.send_text_by_xpath(Element.Login_Username_Xpath_Click,Element.Username)
            SeleniumOperation.send_text_by_xpath(Element.Login_Password_Xpath_Click,Element.Password)
            SeleniumOperation.click_element_by_xpath(Element.Login_Click)
            self.lo.log_to_file(self, "INFO", "Successfully LogIN .")
        except Exception as e:
            self.lo.log_to_file(self, "ERROR", "Exception is : " + str(e))

    def Get_Utilization(self,user):
        SeleniumOperation = SeleniumOperations(self.v_Browser, self.lo)
        try:
            time.sleep(4)
            SeleniumOperation.click_element_by_xpath(Element.Reports_Click)
            SeleniumOperation.click_element_by_xpath(Element.Team_Click)
            time.sleep(8)
            self.Selet_Member(user)
            SeleniumOperation.click_element_by_xpath(Element.Reports_Click)
            T = self.Get_Time()
            SeleniumOperation.click_element_by_xpath(Element.Team_Click)
            SeleniumOperation.click_element_by_xpath(Element.None_Click)
            self.lo.log_to_file(self, "INFO", "Successfully Calculated Utilization for "+user+ " .")
            return T
        except Exception as e:
            self.lo.log_to_file(self, "ERROR", "Exception is : " + str(e))

    def Get_Time(self):
        try:
            time.sleep(3)
            Time = self.v_Browser.find_element_by_xpath(Element.Clock_Click).text
            value = Time.replace(":",".")
            print("Utilization is: "+value)
            # print(Time.replace(":","."))
            return value
        except Exception as e:
            self.lo.log_to_file(self, "ERROR", "Exception is : " + str(e))

    def get_value(self,row_index, column_index):
        file_name = r"C:\Users\Krishnabhashkar.Jha\Desktop\InputSheet.xlsx"
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        v_value = sheet.cell(row=row_index, column=column_index)
        return v_value.value

    def Selet_Member(self,user):
        SeleniumOperation = SeleniumOperations(self.v_Browser, self.lo)
        SeleniumOperation.send_text_by_xpath(Element.Find_Name,user)
        try:
            if user == "adatar":
                SeleniumOperation.click_element_by_xpath(UserCredentials.adatar)
            elif user == "sadake":
                SeleniumOperation.click_element_by_xpath(UserCredentials.sadake)
            elif user == "apawar2":
                SeleniumOperation.click_element_by_xpath(UserCredentials.apawar2)
            elif user == "ainamdar":
                SeleniumOperation.click_element_by_xpath(UserCredentials.ainamdar)
            elif user == "Ajay Choudhary":
                SeleniumOperation.click_element_by_xpath(UserCredentials.AjayChoudhary)
            elif user == "mkamble":
                SeleniumOperation.click_element_by_xpath(UserCredentials.mkamble)
            elif user == "averma":
                SeleniumOperation.click_element_by_xpath(UserCredentials.averma)
            elif user == "amane":
                SeleniumOperation.click_element_by_xpath(UserCredentials.amane)
            elif user == "srathore":
                SeleniumOperation.click_element_by_xpath(UserCredentials.srathore)
            elif user == "nmahala":
                SeleniumOperation.click_element_by_xpath(UserCredentials.nmahala)
            elif user == "sswami":
                SeleniumOperation.click_element_by_xpath(UserCredentials.sswami)
            elif user == "rchoudhari":
                SeleniumOperation.click_element_by_xpath(UserCredentials.rchoudhari)
            elif user == "vtemkar":
                SeleniumOperation.click_element_by_xpath(UserCredentials.vtemkar)
            elif user == "asarode":
                SeleniumOperation.click_element_by_xpath(UserCredentials.asarode)
            elif user == "rkute":
                SeleniumOperation.click_element_by_xpath(UserCredentials.rkute)
            elif user == "pshende":
                SeleniumOperation.click_element_by_xpath(UserCredentials.pshende)
            elif user == "njalkote":
                SeleniumOperation.click_element_by_xpath(UserCredentials.njalkote)
            elif user == "syadav":
                SeleniumOperation.click_element_by_xpath(UserCredentials.syadav)
            elif user == "ssharma":
                SeleniumOperation.click_element_by_xpath(UserCredentials.ssharma)
            elif user == "rsingh":
                SeleniumOperation.click_element_by_xpath(UserCredentials.rsingh)
            elif user == "praut":
                SeleniumOperation.click_element_by_xpath(UserCredentials.praut)
            elif user == "sjakkanawar":
                SeleniumOperation.click_element_by_xpath(UserCredentials.sjakkanawar)
            elif user == "smishra2":
                SeleniumOperation.click_element_by_xpath(UserCredentials.smishra2)
            elif user == "vsingh":
                SeleniumOperation.click_element_by_xpath(UserCredentials.vsingh)
            elif user == "ymaurya":
                SeleniumOperation.click_element_by_xpath(UserCredentials.ymaurya)
            elif user == "skumari2":
                SeleniumOperation.click_element_by_xpath(UserCredentials.skumari2)
            elif user == "ktiwari":
                SeleniumOperation.click_element_by_xpath(UserCredentials.ktiwari)
            elif user == "agurav":
                SeleniumOperation.click_element_by_xpath(UserCredentials.agurav)
            elif user == "sshetty":
                SeleniumOperation.click_element_by_xpath(UserCredentials.sshetty)
            elif user == "gpawar3":
                SeleniumOperation.click_element_by_xpath(UserCredentials.gpawar3)
            elif user == "udhotre":
                SeleniumOperation.click_element_by_xpath(UserCredentials.udhotre)
            elif user == "asalunkhe":
                SeleniumOperation.click_element_by_xpath(UserCredentials.asalunkhe)
            elif user == "ashinde":
                SeleniumOperation.click_element_by_xpath(UserCredentials.ashinde)
            self.lo.log_to_file(self, "INFO", "selected " + user + " for Utilization .")
        except Exception as e:
            self.lo.log_to_file(self, "ERROR", "Exception is : " + str(e))


    def Logout(self):
        SeleniumOperation = SeleniumOperations(self.v_Browser, self.lo)
        try:
            SeleniumOperation.click_element_by_xpath(Element.Sc_Click)
            SeleniumOperation.click_element_by_xpath(Element.LogOut_Click)
            self.lo.log_to_file(self, "INFO", "Successfully LogOut .")
        except Exception as e:
            self.lo.log_to_file(self, "ERROR", "Exception is : " + str(e))

    def Excel_Update(self,row,col,Time):
        v_filePath = r"C:\Users\Krishnabhashkar.Jha\Desktop\InputSheet.xlsx"
        v_row = row
        v_col = col
        ExcelOperations.set_value_to_cell(v_filePath, v_row, v_col,Time)

    def Set_Week(self,week):
        try:
            SeleniumOperation = SeleniumOperations(self.v_Browser, self.lo)
            time.sleep(10)
            SeleniumOperation.click_element_by_xpath(Element.Reports_Click)
            SeleniumOperation.click_element_by_xpath(Element.Duration_Click)
            if week == "Today":
                SeleniumOperation.click_element_by_xpath(Element.Today_Click)
            elif week == "Yesterday":
                SeleniumOperation.click_element_by_xpath(Element.Yesterday_Click)
            elif week == "Thisweek":
                SeleniumOperation.click_element_by_xpath(Element.Thisweek_Click)
            elif week == "Last Week":
                SeleniumOperation.click_element_by_xpath(Element.Lastweek_Click)
            elif week == "Thismonth":
                SeleniumOperation.click_element_by_xpath(Element.Thismonth_Click)
            elif week == "Last Month":
                SeleniumOperation.click_element_by_xpath(Element.Lastmonth_Click)
            elif week == "Thisyear":
                SeleniumOperation.click_element_by_xpath(Element.Thisyear_Click)
            elif week == "Lastyear":
                SeleniumOperation.click_element_by_xpath(Element.Lastyear_Click)
            time.sleep(10)
        except Exception as e:
            print(e)
            self.lo.log_to_file(self, "ERROR", "Exception is : " + str(e))

    def set_value_to_cell(self,row_index, column_index, key):
        file_name = r"C:\Users\Krishnabhashkar.Jha\Desktop\InputSheet.xlsx"
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        sheet.cell(row_index, column_index, key)
        wb.save(filename=file_name)

    def Utilization_calculator(self):
        file_name = r"C:\Users\Krishnabhashkar.Jha\Desktop\InputSheet.xlsx"
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        working_Days = sheet.cell(row=1, column=5).value  # ;print(working_Days)
        working_hours = working_Days * (7.5)  # ;print(working_hours)
        for i in range(4, sheet.max_row + 1, 1):
            v_value = sheet.cell(row=i, column=3).value
            utilization_calculation = (float(v_value) / float(working_hours)) * int((100));
            print(utilization_calculation)
            # time.sleep(1)
            self.set_value_to_cell(i, 4,utilization_calculation)
            # print(i)



